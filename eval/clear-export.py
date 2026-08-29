"""
Pull the CLEAR corpus out of its xlsx into JSONL for `clear-correlate.mjs`.

The corpus is not vendored — CC BY-NC-SA 4.0, and 3.3 MB. Fetch it from
github.com/scrosseye/CLEAR-Corpus first. Needs openpyxl; a venv is fine.

Usage: python clear-export.py CLEAR_corpus_final.xlsx clear.jsonl
"""

import json
import sys

from openpyxl import load_workbook

FORMULA_COLUMNS = {
    "fk": "Flesch-Kincaid-Grade-Level",
    "fre": "Flesch-Reading-Ease",
    "ari": "Automated Readability Index",
    "smog": "SMOG Readability",
    "dc": "New Dale-Chall Readability Formula",
    "carec": "CAREC",
}


def main(xlsx: str, out: str) -> None:
    ws = load_workbook(xlsx, read_only=True, data_only=True)["Data"]
    rows = ws.iter_rows(values_only=True)
    ix = {h: i for i, h in enumerate(next(rows))}

    written = 0
    with open(out, "w") as f:
        for r in rows:
            excerpt, bt = r[ix["Excerpt"]], r[ix["BT_easiness"]]
            if not excerpt or bt is None:
                continue
            rec = {"id": r[ix["ID"]], "text": str(excerpt), "bt": float(bt)}
            rec.update({k: r[ix[col]] for k, col in FORMULA_COLUMNS.items()})
            f.write(json.dumps(rec) + "\n")
            written += 1
    print(f"exported {written} rows to {out}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2])
